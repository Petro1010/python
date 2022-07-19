from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Read workbook and select sheet
wb = load_workbook('barchart.xlsx')
sheet = wb['Report']

# Active rows and columns
min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

for col in range(min_column + 1, max_column + 1):
    colLetter = get_column_letter(col)
    sheet[f'{colLetter}{max_row + 1}'] = f'=SUM({colLetter}{min_row + 1}:{colLetter}{max_row})'  # set the cell formula
    sheet[f'{colLetter}{max_row + 1}'].style = 'Currency'

wb.save('barchart.xlsx')